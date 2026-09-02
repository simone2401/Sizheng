#!/usr/bin/env python3
import json
import hashlib
from pathlib import Path
from typing import Any

import openpyxl


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def sheet_to_records(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else f"col_{i+1}" for i, c in enumerate(rows[0])]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {header[i]: clean_value(row[i] if i < len(row) else None) for i in range(len(header))}
        if any(v is not None for v in record.values()):
            records.append(record)
    return records


def meta_sheet_to_record(ws) -> list[dict[str, Any]]:
    record: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        key = clean_value(row[0] if len(row) > 0 else None)
        value = clean_value(row[1] if len(row) > 1 else None)
        if key is not None:
            record[str(key)] = value
    return [record] if record else []


def unique_by_key(records: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    seen = set()
    result = []
    for record in records:
        value = record.get(key)
        if value is None or value in seen:
            continue
        seen.add(value)
        result.append(record)
    return result


def build_entities(sheets: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    sections = sheets.get("sections", [])
    chapters = unique_by_key(
        [{"chapter_id": row.get("chapter_id"), "chapter_title": row.get("chapter_title")} for row in sections],
        "chapter_id",
    )

    l3_tags = sheets.get("ideology_tag_taxonomy", [])
    level_1 = unique_by_key(
        [{"l1_code": row.get("l1_code"), "l1_label": row.get("l1_label")} for row in l3_tags],
        "l1_code",
    )
    level_2 = unique_by_key(
        [
            {
                "l2_code": row.get("l2_code"),
                "l2_label": row.get("l2_label"),
                "l1_code": row.get("l1_code"),
            }
            for row in l3_tags
        ],
        "l2_code",
    )
    level_3 = unique_by_key(
        [
            {
                "l3_code": row.get("l3_code"),
                "l3_label": row.get("l3_label"),
                "l2_code": row.get("l2_code"),
            }
            for row in l3_tags
        ],
        "l3_code",
    )

    return {
        "textbook": sheets.get("meta", []),
        "chapters": chapters,
        "sections": sections,
        "curriculum_standards": sheets.get("curriculum_standards", []),
        "knowledge_points": sheets.get("knowledge_points", []),
        "textbook_chunks": sheets.get("textbook_original_chunks", []),
        "ideology_paragraphs": sheets.get("ideology_paragraphs", []),
        "ideology_tags": {"level_1": level_1, "level_2": level_2, "level_3": level_3},
    }


def build_relations(sheets: dict[str, list[dict[str, Any]]]) -> dict[str, list[dict[str, Any]]]:
    sections = sheets.get("sections", [])
    taxonomy = sheets.get("ideology_tag_taxonomy", [])

    relations = {
        "chapter_has_section": [
            {"chapter_id": row.get("chapter_id"), "section_id": row.get("section_id")} for row in sections
        ],
        "section_has_standard": [
            {"section_id": row.get("section_id"), "standard_item_id": row.get("standard_item_id")}
            for row in sheets.get("curriculum_standards", [])
        ],
        "section_has_knowledge_point": [
            {"section_id": row.get("section_id"), "knowledge_point_id": row.get("knowledge_point_id")}
            for row in sheets.get("knowledge_points", [])
        ],
        "section_has_chunk": [
            {"section_id": row.get("section_id"), "chunk_id": row.get("chunk_id")}
            for row in sheets.get("textbook_original_chunks", [])
        ],
        "section_has_ideology_paragraph": [
            {"section_id": row.get("section_id"), "paragraph_id": row.get("paragraph_id")}
            for row in sheets.get("ideology_paragraphs", [])
        ],
        "tag_l1_has_l2": unique_by_key(
            [{"l1_code": row.get("l1_code"), "l2_code": row.get("l2_code")} for row in taxonomy], "l2_code"
        ),
        "tag_l2_has_l3": unique_by_key(
            [{"l2_code": row.get("l2_code"), "l3_code": row.get("l3_code")} for row in taxonomy], "l3_code"
        ),
        "paragraph_tagged_with_l3": [],
    }

    for row in sheets.get("ideology_paragraphs", []):
        paragraph_id = row.get("paragraph_id")
        codes = row.get("ideology_l3_codes")
        if paragraph_id is None or codes is None:
            continue
        for code in str(codes).split(";"):
            tag_code = code.strip()
            if tag_code:
                relations["paragraph_tagged_with_l3"].append(
                    {"paragraph_id": paragraph_id, "l3_code": tag_code}
                )

    for key in relations:
        relations[key] = [
            row
            for row in relations[key]
            if all(v is not None and not (isinstance(v, str) and not v.strip()) for v in row.values())
        ]

    return relations


def build_summary(source_file: str, sheets: dict[str, list[dict[str, Any]]], entities: dict[str, Any], relations: dict[str, Any]) -> dict[str, Any]:
    return {
        "source_file": source_file,
        "sheet_count": len(sheets),
        "sheets": list(sheets.keys()),
        "sheet_row_counts": {name: len(rows) for name, rows in sheets.items()},
        "entity_counts": {
            "textbook": len(entities["textbook"]),
            "chapters": len(entities["chapters"]),
            "sections": len(entities["sections"]),
            "curriculum_standards": len(entities["curriculum_standards"]),
            "knowledge_points": len(entities["knowledge_points"]),
            "textbook_chunks": len(entities["textbook_chunks"]),
            "ideology_paragraphs": len(entities["ideology_paragraphs"]),
            "ideology_tags": {
                "level_1": len(entities["ideology_tags"]["level_1"]),
                "level_2": len(entities["ideology_tags"]["level_2"]),
                "level_3": len(entities["ideology_tags"]["level_3"]),
            },
        },
        "relation_counts": {name: len(rows) for name, rows in relations.items()},
    }


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_topic_type(cognitive_level: Any) -> str:
    text = str(cognitive_level or "")
    if any(token in text for token in ["操作", "实验", "测量"]):
        return "PROCEDURAL"
    if "表达" in text or "语言" in text:
        return "LANGUAGE"
    return "CONCEPTUAL"


def mk_topic_id(raw_id: Any) -> str:
    source = str(raw_id or "")
    digest = hashlib.sha1(source.encode("utf-8")).hexdigest()[:12]
    return f"mt_{digest}"


def build_marble_graph(sheets: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    meta = sheets.get("meta", [{}])[0] if sheets.get("meta") else {}
    sections = sheets.get("sections", [])
    standards = sheets.get("curriculum_standards", [])
    points = sheets.get("knowledge_points", [])

    section_to_chapter_title = {row.get("section_id"): row.get("chapter_title") for row in sections}
    section_to_standards: dict[str, list[str]] = {}
    for std in standards:
        sid = std.get("section_id")
        code = std.get("code")
        if sid and code:
            section_to_standards.setdefault(str(sid), []).append(str(code))

    topics: list[dict[str, Any]] = []
    kp_to_topic: dict[str, str] = {}
    section_kps: dict[str, list[str]] = {}

    for kp in points:
        kp_id = kp.get("knowledge_point_id")
        section_id = kp.get("section_id")
        if not kp_id or not section_id:
            continue
        topic_id = mk_topic_id(kp_id)
        kp_to_topic[str(kp_id)] = topic_id
        section_kps.setdefault(str(section_id), []).append(topic_id)

        topics.append(
            {
                "id": topic_id,
                "type": normalize_topic_type(kp.get("cognitive_level")),
                "subject": meta.get("subject") or "物理",
                "domain": section_to_chapter_title.get(section_id),
                "name": kp.get("title"),
                "description": kp.get("summary") or "",
                "ageRangeStart": None,
                "ageRangeEnd": None,
                "centrality": None,
                "evidence": [str(kp.get("cognitive_level"))] if kp.get("cognitive_level") else [],
                "assessmentPrompt": None,
                "standards": section_to_standards.get(str(section_id), []),
                "source": {
                    "knowledge_point_id": kp_id,
                    "section_id": section_id,
                    "printed_pages": kp.get("printed_pages"),
                },
            }
        )

    dependencies: list[dict[str, Any]] = []
    for section_id, topic_ids in section_kps.items():
        for i in range(1, len(topic_ids)):
            dependencies.append(
                {
                    "topicId": topic_ids[i],
                    "prerequisiteId": topic_ids[i - 1],
                    "strength": "soft",
                    "reason": f"同一节({section_id})内知识点递进",
                }
            )

    section_ids_in_order = [row.get("section_id") for row in sections if row.get("section_id")]
    for i in range(1, len(section_ids_in_order)):
        prev_sid = str(section_ids_in_order[i - 1])
        cur_sid = str(section_ids_in_order[i])
        prev_topics = section_kps.get(prev_sid, [])
        cur_topics = section_kps.get(cur_sid, [])
        if prev_topics and cur_topics:
            dependencies.append(
                {
                    "topicId": cur_topics[0],
                    "prerequisiteId": prev_topics[-1],
                    "strength": "soft",
                    "reason": "相邻节次的学习衔接",
                }
            )

    curricula = {
        "slug": "cn-pep-physics-8u-2024",
        "country": "CN",
        "name": meta.get("textbook_name") or "教材课程标准",
        "version": str(meta.get("edition") or "unknown"),
        "sourceUrl": None,
        "textIncluded": True,
        "license": "unknown",
        "topicCount": len(standards),
        "topics": [
            {
                "key": str(row.get("standard_item_id") or row.get("code") or ""),
                "code": str(row.get("code") or ""),
                "data": {
                    "sectionId": row.get("section_id"),
                    "sectionCategory": row.get("section_category"),
                    "itemTitle": row.get("item_title"),
                    "itemContent": row.get("item_content"),
                    "printedPage": row.get("printed_page"),
                },
            }
            for row in standards
            if row.get("standard_item_id") or row.get("code")
        ],
    }

    chapter_titles = []
    seen_chapter = set()
    for row in sections:
        title = row.get("chapter_title")
        if title and title not in seen_chapter:
            seen_chapter.add(title)
            chapter_titles.append(title)

    clusters = [
        {
            "subject": meta.get("subject") or "物理",
            "domain": chapter,
            "ageRangeStart": 13,
            "summary": f"{chapter}：涵盖本章主要知识点与应用能力要求。",
        }
        for chapter in chapter_titles
    ]

    topics_doc = {"version": "v1-local", "topicCount": len(topics), "topics": topics}
    dependencies_doc = {
        "version": "v1-local",
        "note": "Auto-generated from textbook knowledge points",
        "edgeCount": len(dependencies),
        "dependencies": dependencies,
    }
    standards_doc = {
        "note": "Mapped from workbook curriculum_standards sheet",
        "codesOnlySources": [],
        "curriculumCount": 1,
        "curricula": [curricula],
    }
    clusters_doc = {"version": "v1-local", "clusterCount": len(clusters), "clusters": clusters}

    return {
        "topics": topics_doc,
        "dependencies": dependencies_doc,
        "curriculum_standards": standards_doc,
        "clusters": clusters_doc,
    }


def build_marble_manifest(marble_dir: Path, marble_graph: dict[str, Any]) -> dict[str, Any]:
    files = {
        "topics.json": marble_graph["topics"],
        "dependencies.json": marble_graph["dependencies"],
        "curriculum-standards.json": marble_graph["curriculum_standards"],
        "clusters.json": marble_graph["clusters"],
    }

    manifest_files = {}
    for name in files:
        path = marble_dir / name
        payload = path.read_bytes()
        manifest_files[name] = {"bytes": len(payload), "sha256": hashlib.sha256(payload).hexdigest()}

    by_subject: dict[str, int] = {}
    for topic in marble_graph["topics"]["topics"]:
        subject = str(topic.get("subject") or "Unknown")
        by_subject[subject] = by_subject.get(subject, 0) + 1

    return {
        "version": "v1-local",
        "counts": {
            "topics": marble_graph["topics"]["topicCount"],
            "dependencies": marble_graph["dependencies"]["edgeCount"],
            "clusters": marble_graph["clusters"]["clusterCount"],
            "curricula": marble_graph["curriculum_standards"]["curriculumCount"],
        },
        "subjects": by_subject,
        "files": manifest_files,
    }


def convert_workbook(path: Path, output_root: Path) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        if name == "meta":
            sheets[name] = meta_sheet_to_record(ws)
        else:
            sheets[name] = sheet_to_records(ws)

    entities = build_entities(sheets)
    relations = build_relations(sheets)
    summary = build_summary(path.name, sheets, entities, relations)
    graph = {"entities": entities, "relations": relations, "summary": summary}

    workbook_dir = output_root / path.stem
    raw_dir = workbook_dir / "raw_sheets"
    marble_dir = workbook_dir / "marble"

    for sheet_name, rows in sheets.items():
        write_json(raw_dir / f"{sheet_name}.json", rows)

    write_json(workbook_dir / "entities.json", entities)
    write_json(workbook_dir / "relations.json", relations)
    write_json(workbook_dir / "summary.json", summary)
    write_json(workbook_dir / "graph.json", graph)

    marble_graph = build_marble_graph(sheets)
    write_json(marble_dir / "topics.json", marble_graph["topics"])
    write_json(marble_dir / "dependencies.json", marble_graph["dependencies"])
    write_json(marble_dir / "curriculum-standards.json", marble_graph["curriculum_standards"])
    write_json(marble_dir / "clusters.json", marble_graph["clusters"])
    write_json(marble_dir / "manifest.json", build_marble_manifest(marble_dir, marble_graph))


def main() -> None:
    base = Path(".")
    output_root = base / "json_output"
    excel_files = sorted(base.glob("*.xlsx")) + sorted(base.glob("*.xls"))

    if not excel_files:
        print("未找到 Excel 文件（.xlsx/.xls）。")
        return

    for workbook in excel_files:
        convert_workbook(workbook, output_root)
        print(f"已转换: {workbook.name} -> {output_root / workbook.stem}")


if __name__ == "__main__":
    main()
